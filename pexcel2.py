#!/usr/bin/env python

#-*- coding: UTF-8 -*-
import sys
import os
from openpyxl import load_workbook
#from tkFileDialog import askopenfilename
from tkinter import filedialog
import tkinter as tk


def sel_file():
    global filename
    #filename = askopenfilename()
    filename = filedialog.askopenfilename()    

def append_text(sheet,rr, cc,text):
    wb = load_workbook(filename)
    ws = wb[sheet]
    print ("wroking sheet is:"+ws.title)
    print ("write row=",rr," column=",cc," value=",text)
    ws.cell(row=rr,column=cc).value=text
    wb.save(filename)
    ws=wb['sheet1']
    for ir in range(1,6):
        for ic in range(1,3):
            print (ws.cell(row=ir, column=ic).value)

def update_text():
    sheet=var1.get()
    rr=int(var2.get())
    cc=int(var3.get())
    str1=var4.get()
    append_text(sheet,rr,cc,str1)
    text_output.delete(0.0,'end')
    text_output.insert(0.0,get_text())
    #text_output.config(text=get_text())
    #var1.set(' ')
    #var2.set(' ')
    #var3.set(' ')
    #var4.set(' ')

def get_text():
    wb = load_workbook(filename)
    ws=wb['sheet1']
    text=""
    for ir in range(1,6):
        for ic in range(1,3):
            str1 =ws.cell(row=ir, column=ic).value
            text = str(text) +"\n"+ str(str1)
    return text


root = tk.Tk()
root.title("Excel file")
var1 = tk.StringVar(value="sheet")
text_input = tk.Entry(root,textvariable=var1,width=36, bd=5).grid(row=1,sticky='n')
var2 = tk.StringVar(value="row")
text_input = tk.Entry(root,textvariable=var2,width=36, bd=5).grid(row=2,sticky='n')
var3 = tk.StringVar(value="column")
text_input = tk.Entry(root,textvariable=var3,width=36, bd=5).grid(row=3,sticky='n')
var4 = tk.StringVar(value="value")
text_input = tk.Entry(root,textvariable=var4,width=36, bd=5).grid(row=4,sticky='n')

tk.Button(root, text="dump excel file", command=update_text).grid(row=5,column=0,sticky='e')
root.bind('<Return>', lambda event:update_text())

tk.Button(root, text="select file", command=sel_file).grid(row=5,column=0,sticky='w')
root.bind('<Return>', lambda event:sel_file())


text_output=tk.Text(root,width=50,height=10,wrap='none')
text_output.grid(row=6,sticky='news')
s=tk.Scrollbar(root)
s.grid(row=6,column=1,sticky='news')
s.config(command=text_output.yview)

sd=tk.Scrollbar(root,orient= 'horizontal')
sd.grid(row=7,sticky='news')
sd.config(command=text_output.xview)

text_output.config(xscrollcommand=sd.set,yscrollcommand=s.set)


#wb = load_workbook(filename)
#print (wb.get_sheet_names())

#for sheet in wb:
#     print(sheet.title)
     



#ws['A4'] = 40
#print wb.get_sheet_names()
#wb.save('test.xlsx')

##for rr in range(1,6):
##	for col in range(1,3):
##	   print (ws.cell(row=rr, column=col).value)
root.mainloop()	   

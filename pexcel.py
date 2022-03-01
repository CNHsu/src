#!/usr/bin/env python

#-*- coding: UTF-8 -*-
import sys
from openpyxl import load_workbook

wb = load_workbook('bb.xlsx')
print (wb.get_sheet_names())

for sheet in wb:
     print(sheet.title)
     
ws = wb['PE cyc=1.4K_60K before_baking']
print (ws.title)
#ws.title = "sheet2" # modify sheet name
print ("wroking sheet is:"+ws.title)


#ws['A4'] = 40
#print wb.get_sheet_names()
#wb.save('test.xlsx')

for rr in range(1,6):
	for col in range(1,3):
	   print (ws.cell(row=rr, column=col).value)